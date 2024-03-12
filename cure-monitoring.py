# Raspi Aushaerteueberwachung

# Funktion:
# Knopf 1s drücken, um die Messung zu starten (ab jetzt werden Temperaturdaten eingelesen und über das Kinetikmodell der Aushärtegrad bestimmt)
# Knopf erneut für 1s drücken, um die Messung wieder zu stoppen (die Daten werden in einer Exceltabelle mit den Spalten Zeit[s], Temperatur[K] und alpha gespeichert und im selben Ordner abespeichert, in dem sich dieses Programm befindet)

import numpy as np
import time
import digitalio
import board
import adafruit_max31865
from math import isnan
import RPi.GPIO as GPIO     # ('#' zum Testen am PC)
import openpyxl

# Kinetikmodell nach Grindling:
def grindlingModell(alpha, T):

    K_2diffT_g, E_1, E_2, c_1, c_2, deltaTg, A_1, A_2, n_1, n_2, m = (1.88963840e-02, 7.74862316e+04, 5.26859969e+04, 4.56550329e-01, 1.44188898e+00, 2.78375900e+02, 1.75745854e+08, 2.25672390e+05, 3.07641009e+00, 1.31031748e+00, 6.25391337e-01)
    G_1 = 0.718172483
    G_2 = 2.375009844
    R = 8.31446     # ideale Gaskonstante

    T_g = (-37.195+273.15) * np.exp((G_1*alpha)/(G_2-alpha))
    E_1diff = R*pow(T_g,2)*(c_1/c_2)
    E_2diff = R*(c_1*c_2*pow((T_g+deltaTg),2))/pow((c_2+deltaTg),2)

    if T < T_g:
        K_2diff = K_2diffT_g * np.exp((-E_1diff/R)*(1/T-1/T_g))
    elif T <= (T_g+deltaTg):
        K_2diff = K_2diffT_g * np.exp((c_1*(T-T_g))/(c_2+T-T_g))
    else:
        K_2diff = K_2diffT_g * np.exp(c_1*(deltaTg)/(c_2+deltaTg))*np.exp((-E_2diff/R)*(1/T-1/(T_g+deltaTg)))

    K_1 = A_1*np.exp(-E_1/(R*T))
    K_2 = A_2*np.exp(-E_2/(R*T))
    K_eff = 1/(1/K_2diff+1/K_2)

    deltaAlphaDeltaZeit = (K_1*pow((1-alpha),n_1)+K_eff*pow(alpha,m)*pow((1-alpha),n_2))

    return deltaAlphaDeltaZeit

def exportDataToExcel(data, name_exceldatei, name_sheet):   # schreibt eine list of lists in eine Exceltabelle gemäß listOfLists[zeile][spalte].

    workbook = openpyxl.Workbook()
    workbook.create_sheet(name_sheet)
    sheet = workbook[name_sheet]

    row_max = len(data)
    for row in range(row_max):
        col_max = len(data[row])
        for col in range(col_max):
            sheet.cell(row=row+1, column=col+1).value = data[row][col]

    workbook.save(filename=str(name_exceldatei))
    workbook.close()

def readTemperature():      # liest die Temperatur des Sensors aus. Falls der Wert 'nan' entspricht, wird ein neuer Messwert angefordert, bis eine gültige Zahl vorliegt
    temp = sensor.temperature
#    temp = 100  # Testtemperatur
    if isnan(temp):
        return(readTemperature())
    else:
        return(temp)

def giveTemperatureValue():     # nimmt 7 Temperaturwerte auf, entfernt die 2 höchsten und die 2 niedrigsten Werten und gibt den Durchschnitt der 3 verbleibenden Werte zurück
    temperatures = []
    for i in range(1,7):
        temp = readTemperature()
        temperatures.append(temp)
    temperatures.sort()
    trimmed = temperatures[2:-2]        # höchste und niedrigste zwei Werte werden entfernt
    average = sum(trimmed)/len(trimmed) # Durchschnittswert der Messwerte berechnen
    return(average)




button_pushed = False   # ('#' für Testen am PC)
#button_pushed = True   # ('#' für Nutzung am Raspi)
button_last_time_pressed = 0
alpha = 0.
datenset = []
clock = time.time()
startzeit = time.time()
GPIO.setup(17, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)         # ()'#' zum Testen am PC)
# setup PT1000 mit MAX31865
cs = digitalio.DigitalInOut(board.D5)
spi = board.SPI()
sensor = adafruit_max31865.MAX31865(spi,cs,rtd_nominal=1000,ref_resistor=4300.0,wires=2)


pause_between_measurements = 0.05    # beugt zudem Prellen beim betätigen des Knopfes vor

print('Script läuft. Drücke Strg-C um abzubrechen.')

while True:
#for i in range(1,1000):    # (zum Testen am PC anstelle 'while True')
    if (time.time()-button_last_time_pressed > 5):  # ein Knopfdruck wird erst wieder 5s nach der letzten Betätigung registriert
        if GPIO.input(17) == GPIO.HIGH: # ('False' zum Testen am PC)
            
            if button_pushed == True:   # toggle button
                button_pushed = False
            else:
                button_pushed = True

            button_last_time_pressed = time.time()
            if (button_pushed == True):
                clock = time.time()     # Startzeitpunkt der Messung setzen
                startzeit = clock
                print ('Messung gestartet.')
            else:                       # Messung Abschließen
                zeit = time.localtime(time.time())
                zeit_str = str(zeit.tm_year)+'-'+str(zeit.tm_mon)+'-'+str(zeit.tm_mday)+' '+str(zeit.tm_hour)+'-'+str(zeit.tm_min)+'-'+str(zeit.tm_sec)
                dateiname = zeit_str+'Aushaerteverlauf.xlsx'
                exportDataToExcel(datenset,dateiname,zeit_str)
                print ('Messung beendet und Daten in Exceltabelle gespeichert. Drücke Strg-C um das Skript zu beenden oder betätige den Knopf, um eine neue Messung zu starten.')

    if (button_pushed == True):
        temp = giveTemperatureValue()+273.15    # aktuelle Temperatur in K anfordern
        clock_new = time.time()
        delta_t = clock_new - clock
        alpha = alpha+(grindlingModell(alpha,temp) * delta_t)
        new_set = []
        new_set.append(clock_new-startzeit)
        new_set.append(temp)
        new_set.append(alpha)
        datenset.append(new_set)
        clock = clock_new
        print('Zeit[s], Temperatur[K], alpha:', new_set)
    else:
        t = sensor.temperature
        print('Messung läuft nicht.')
        print('Temperatur: {0:0.3F}*C'.format(t))


    time.sleep(pause_between_measurements)


